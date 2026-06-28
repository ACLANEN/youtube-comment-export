"""
YouTube Comment Research Tool — by Fynn
"""
import os
import csv
import io
import json
import uuid
import threading
import time
from datetime import datetime, timedelta

import requests
import xml.etree.ElementTree as ET
import glob as _glob

import logging
logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

from flask import Flask, render_template, request, jsonify, send_file, redirect

# ── 可选依赖（不阻塞启动）──
try:
    import yt_dlp
    _HAS_YTDLP = True
except ImportError:
    _HAS_YTDLP = False
    logger.warning("yt-dlp 未安装，字幕方案 B 不可用")

try:
    from deep_translator import GoogleTranslator
    _HAS_TRANSLATOR = True
except ImportError:
    _HAS_TRANSLATOR = False
    logger.warning("deep-translator 未安装，翻译功能不可用")

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# ── 请求耗时日志 ──
@app.before_request
def _log_request_start():
    request._start_time = time.time()

@app.after_request
def _log_request_end(response):
    if hasattr(request, '_start_time'):
        elapsed = time.time() - request._start_time
        if elapsed > 3.0:
            logger.warning(f"slow request: {request.method} {request.path} took {elapsed:.1f}s")
    return response

# ── 统一错误响应 ──
def _err(msg: str, code: int = 400):
    return jsonify({"ok": False, "error": msg}), code

# ── 配置（仅从环境变量读取，无默认值暴露）──
API_KEY = os.getenv("YOUTUBE_API_KEY")
BASE_URL = os.getenv("YOUTUBE_BASE_URL", "https://www.googleapis.com/youtube/v3")
CAPTION_PROXY = os.getenv("CAPTION_PROXY", "")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")

# ── 激活码存储 ──
CODES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "codes.json")
_codes_lock = threading.Lock()

def _load_codes():
    try:
        with open(CODES_FILE, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _save_codes(codes):
    with open(CODES_FILE, "w") as f:
        json.dump(codes, f, ensure_ascii=False, indent=2)

def _get_codes():
    with _codes_lock:
        return _load_codes()

def _update_codes(fn):
    with _codes_lock:
        codes = _load_codes()
        codes = fn(codes)
        _save_codes(codes)
        return codes

# 确保 /data 目录存在
# 启动时初始化
if not os.path.exists(CODES_FILE):
    _save_codes({})


# ── 配额保护：检测到 quotaExceeded 后全局标记，避免浪费后续请求 ──
_quota_exhausted = False

def _youtube_get(path: str, key_override: str = None, retries: int = 3) -> dict:
    global _quota_exhausted
    if _quota_exhausted:
        raise RuntimeError("YouTube API 配额已用完（本次运行期间已检测到）")

    key = key_override or API_KEY
    if not key:
        raise RuntimeError("YOUTUBE_API_KEY 未配置")
    url = f"{BASE_URL}/{path}&key={key}"
    last_err = None
    for attempt in range(retries):
        try:
            resp = requests.get(url, timeout=20)
            if resp.status_code == 403 and "quotaExceeded" in resp.text:
                _quota_exhausted = True
                raise RuntimeError("YouTube API 配额已用完，请明天再试或更换 API Key")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.Timeout as e:
            last_err = e
            if attempt < retries - 1:
                time.sleep(1.5 * (attempt + 1))
        except requests.exceptions.HTTPError as e:
            if resp.status_code == 429 and attempt < retries - 1:
                time.sleep(3 * (attempt + 1))
                continue
            raise
        except requests.exceptions.RequestException as e:
            last_err = e
            if attempt < retries - 1:
                time.sleep(1 * (attempt + 1))
    raise RuntimeError(f"YouTube API 请求失败（重试{retries}次）: {last_err}")


def _format_duration(iso: str) -> str:
    """PT1H23M45S → 1:23:45"""
    import re
    if not iso:
        return ""
    m = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', iso)
    if not m:
        return ""
    h, mm, s = m.groups()
    parts = []
    if h: parts.append(h)
    parts.append(mm or "0")
    sec = s if s else "0"
    parts.append(sec.zfill(2))
    if len(parts) == 3:
        return f"{parts[0]}:{parts[1]}:{parts[2]}"
    return f"{parts[0]}:{parts[1]}"


# ═══════════════════════════════════════
#  激活码 API
# ═══════════════════════════════════════

@app.route("/api/verify-code", methods=["POST"])
def api_verify_code():
    data = request.get_json()
    code = (data.get("code") or "").strip().upper()
    if not code:
        return jsonify({"ok": False, "error": "请输入激活码"}), 400

    codes = _get_codes()
    entry = codes.get(code)

    if not entry:
        return jsonify({"ok": False, "error": "激活码无效"}), 403

    if not entry.get("active", True):
        return jsonify({"ok": False, "error": "激活码已禁用"}), 403

    if entry.get("used", False):
        return jsonify({"ok": False, "error": "激活码已被使用"}), 403

    expires = entry.get("expires_at")
    if expires:
        try:
            exp = datetime.fromisoformat(expires)
            if datetime.now() > exp:
                return jsonify({"ok": False, "error": "激活码已过期"}), 403
        except ValueError:
            pass

    # 标记为已使用
    def _mark_used(codes):
        if code in codes:
            codes[code]["used"] = True
        return codes
    _update_codes(_mark_used)

    return jsonify({"ok": True, "expires_at": expires})


# ═══════════════════════════════════════
#  管理面板
# ═══════════════════════════════════════

@app.route("/admin")
def admin_panel():
    return render_template("admin.html")


@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    data = request.get_json()
    pw = data.get("password", "")
    if pw == ADMIN_PASSWORD:
        return jsonify({"ok": True, "token": "admin_session"})
    return jsonify({"ok": False, "error": "密码错误"}), 403


@app.route("/api/admin/codes", methods=["GET"])
def admin_list_codes():
    codes = _get_codes()
    items = []
    for code, entry in codes.items():
        items.append({
            "code": code,
            "active": entry.get("active", True),
            "used": entry.get("used", False),
            "created_at": entry.get("created_at", ""),
            "expires_at": entry.get("expires_at", ""),
            "note": entry.get("note", ""),
        })
    items.sort(key=lambda x: x["created_at"], reverse=True)
    return jsonify({"codes": items})


@app.route("/api/admin/codes/create", methods=["POST"])
def admin_create_code():
    data = request.get_json()
    note = data.get("note", "").strip()
    days = int(data.get("days", 30))

    # 生成 8 位大写激活码
    code = uuid.uuid4().hex[:8].upper()
    now = datetime.now()
    expires = now + timedelta(days=days)

    def _add(codes):
        codes[code] = {
            "active": True,
            "used": False,
            "created_at": now.isoformat(),
            "expires_at": expires.isoformat(),
            "note": note,
        }
        return codes

    _update_codes(_add)
    return jsonify({"ok": True, "code": code, "expires_at": expires.isoformat()})


@app.route("/api/admin/codes/toggle", methods=["POST"])
def admin_toggle_code():
    data = request.get_json()
    code = (data.get("code") or "").strip().upper()

    def _toggle(codes):
        if code in codes:
            codes[code]["active"] = not codes[code].get("active", True)
        return codes

    _update_codes(_toggle)
    codes = _get_codes()
    entry = codes.get(code, {})
    return jsonify({"ok": True, "active": entry.get("active", True)})


@app.route("/api/admin/codes/delete", methods=["POST"])
def admin_delete_code():
    data = request.get_json()
    code = (data.get("code") or "").strip().upper()

    def _del(codes):
        codes.pop(code, None)
        return codes

    _update_codes(_del)
    return jsonify({"ok": True})


@app.route("/api/admin/codes/update", methods=["POST"])
def admin_update_code():
    data = request.get_json()
    code = (data.get("code") or "").strip().upper()
    days = data.get("days")

    def _upd(codes):
        if code in codes and days:
            now = datetime.now()
            codes[code]["expires_at"] = (now + timedelta(days=int(days))).isoformat()
        return codes

    _update_codes(_upd)
    return jsonify({"ok": True})



@app.route("/api/admin/codes/reset", methods=["POST"])
def admin_reset_code():
    data = request.get_json()
    code = (data.get("code") or "").strip().upper()

    def _reset(codes):
        if code in codes:
            codes[code]["used"] = False
        return codes

    _update_codes(_reset)
    return jsonify({"ok": True})



# ═══════════════════════════════════════
#  字幕提取
# ═══════════════════════════════════════

def _extract_captions(video_id, timeout_sec: int = 30):
    """字幕提取：youtube_transcript_api 优先 → yt-dlp → timedtext API（总超时{timeout_sec}s）"""
    import concurrent.futures
    import html as html_mod

    def _do_extract():
        return _extract_captions_inner(video_id, html_mod)

    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_do_extract)
            return future.result(timeout=timeout_sec)
    except concurrent.futures.TimeoutError:
        logger.warning(f"captions: timed out after {timeout_sec}s for {video_id}")
        return None
    except Exception:
        return None


def _extract_captions_inner(video_id, html_mod):

    segments = None
    lang = "en"

    # === 方案 A: youtube_transcript_api（最可靠，无需 PO token）===
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
        api = YouTubeTranscriptApi()
        transcripts = api.list(video_id)
        # 优先英文手动字幕 → 英文自动字幕 → 其他
        target = None
        for t in transcripts:
            if t.language_code.startswith("en") and not t.is_generated:
                target = t
                break
        if not target:
            for t in transcripts:
                if t.language_code.startswith("en") and t.is_generated:
                    target = t
                    break
        if not target:
            for t in transcripts:
                target = t
                break
        if target:
            fetched = target.fetch()
            lang = target.language_code
            segments = []
            for s in fetched:
                segments.append({
                    "text": html_mod.unescape(s.text),
                    "start": round(s.start, 1),
                    "duration": round(s.duration, 1),
                })
    except Exception as e:
        logger.info(f"captions: transcript_api failed for {video_id}: {e}")

    # === 方案 B: yt-dlp 下载字幕 ===
    if not segments and _HAS_YTDLP:
        ydl_opts = {
            'skip_download': True,
            'writesubtitles': True,
            'writeautomaticsub': True,
            'subtitleslangs': ['en'],
            'subtitlesformat': 'ttml',
            'outtmpl': {'default': f'/tmp/ytdl_{video_id}.%(ext)s'},
            'quiet': True,
            'no_warnings': True,
            'socket_timeout': 20,
            'retries': 2,
        }
        if CAPTION_PROXY:
            ydl_opts['proxy'] = CAPTION_PROXY

        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                ydl.extract_info(f'https://www.youtube.com/watch?v={video_id}', download=True)
            for path in sorted(_glob.glob(f'/tmp/ytdl_{video_id}*.ttml')):
                try:
                    with open(path, 'r') as f:
                        xml_text = f.read()
                    os.remove(path)
                    if len(xml_text) > 50:
                        segs = _parse_ttml(xml_text, html_mod)
                        if segs:
                            segments = segs
                            break
                except Exception:
                    try: os.remove(path)
                    except: pass
        except Exception as e:
            logger.info(f"captions: yt-dlp failed for {video_id}: {e}")

    # === 方案 C: timedtext API 直连 ===
    if not segments:
        UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        for ln in ["en", "en-US", "en-GB"]:
            try:
                proxies = {"http": CAPTION_PROXY, "https": CAPTION_PROXY} if CAPTION_PROXY else None
                r = requests.get(
                    f"https://www.youtube.com/api/timedtext?v={video_id}&lang={ln}",
                    headers={"User-Agent": UA}, timeout=10, proxies=proxies
                )
                if r.status_code == 200 and len(r.text) > 50:
                    segments = _parse_ttml(r.text, html_mod)
                    lang = ln
                    break
            except Exception as e:
                logger.info(f"captions: timedtext failed for {video_id} ({ln}): {e}")

    if not segments:
        logger.warning(f"captions: all 3 methods failed for {video_id}")
        return None

    return {
        "video_id": video_id,
        "language": lang,
        "segments": segments,
        "total": len(segments),
    }


def _parse_ttml(xml_text, html_mod):
    try:
        root = ET.fromstring(xml_text)
        segments = []
        for p in root.iter("{http://www.w3.org/ns/ttml}p"):
            begin = p.attrib.get("begin", "0")
            dur = p.attrib.get("dur", "1")
            text = "".join(p.itertext()).strip()
            if text:
                s = _parse_ttml_time(begin)
                d = _parse_ttml_time(dur)
                segments.append({"text": html_mod.unescape(text), "start": round(s, 1), "duration": round(d, 1)})
        return segments
    except ET.ParseError as e:
        logger.info(f"ttml parse error: {e}")
        return []
    except Exception as e:
        logger.error(f"ttml parse unexpected error: {e}")
        return []


def _parse_ttml_time(t):
    t = t.strip()
    if ":" in t:
        parts = t.split(":")
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + float(parts[1])
    return float(t) if t else 0


@app.route("/api/captions")
def api_captions():
    video_id = request.args.get("video_id", "").strip()
    if not video_id:
        return jsonify({"error": "缺少 video_id"}), 400
    result = _extract_captions(video_id)
    if result is None:
        return jsonify({"error": "字幕获取失败（视频可能无字幕或被封锁）"}), 404
    return jsonify({
        "video_id": video_id,
        "language": result["language"],
        "segments": result["segments"],
        "total_segments": result["total"],
    })


@app.route("/api/export/transcript", methods=["POST"])
def api_export_transcript():
    data = request.get_json()
    video_ids = data.get("video_ids", [])
    fmt = data.get("format", "csv")
    structure = data.get("structure", "flat")
    if not video_ids:
        return jsonify({"error": "未选择视频"}), 400

    videos = []
    for vid in video_ids:
        try:
            vdata = _youtube_get(f"videos?part=snippet&id={vid}")
        except Exception:
            continue
        items = vdata.get("items", [])
        if not items:
            continue
        title = items[0]["snippet"]["title"]
        result = _extract_captions(vid)
        if not result:
            continue
        segs = []
        for seg in result["segments"]:
            ss = int(seg["start"])
            mm = ss // 60; ss2 = ss % 60
            ts = f"{mm:02d}:{ss2:02d}"
            link = f"https://www.youtube.com/watch?v={vid}&t={ss}"
            segs.append({"timestamp": ts, "duration": round(seg["duration"], 1), "text": seg["text"], "link": link})
        videos.append({"video_id": vid, "title": title, "segments": segs})

    if not videos:
        return jsonify({"error": "所选视频无字幕"}), 404

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"transcript_{date_str}"

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook; from openpyxl.styles import Font, PatternFill
        except ImportError:
            return jsonify({"error": "需要openpyxl"}), 500
        wb = Workbook(); hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        if structure == "per_video":
            first = True
            for v in videos:
                name = v["title"][:31].replace("/","-")
                ws = wb.active if first else wb.create_sheet(name); first = False
                ws.title = name
                for col, h in enumerate(["Timestamp", "Duration", "Transcript", "Video Link"], 1):
                    cell = ws.cell(row=1, column=col, value=h); cell.font = hf; cell.fill = hfill
                for r, s in enumerate(v["segments"], 2):
                    for ci, val in enumerate([s["timestamp"], s["duration"], s["text"], s["link"]], 1):
                        ws.cell(row=r, column=ci, value=val)
                ws.column_dimensions["A"].width = 10; ws.column_dimensions["B"].width = 8
                ws.column_dimensions["C"].width = 60; ws.column_dimensions["D"].width = 40
        else:
            ws = wb.active; ws.title = "字幕"
            for col, h in enumerate(["Video Title", "Video ID", "Timestamp", "Duration", "Transcript", "Video Link"], 1):
                cell = ws.cell(row=1, column=col, value=h); cell.font = hf; cell.fill = hfill
            r = 2
            for v in videos:
                for s in v["segments"]:
                    for ci, val in enumerate([v["title"], v["video_id"], s["timestamp"], s["duration"], s["text"], s["link"]], 1):
                        ws.cell(row=r, column=ci, value=val)
                    r += 1
            for col, w in enumerate([40, 14, 10, 8, 60, 40], 1):
                ws.column_dimensions[chr(64+col)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"{filename}.xlsx")
    elif fmt == "json":
        if structure == "per_video":
            result = {"exported_at": datetime.now().isoformat(), "videos": []}
            for v in videos:
                result["videos"].append({"title": v["title"], "video_id": v["video_id"], "segments": v["segments"]})
        else:
            result = {"exported_at": datetime.now().isoformat(), "transcripts": []}
            for v in videos:
                for s in v["segments"]:
                    result["transcripts"].append({"video_title": v["title"], "video_id": v["video_id"],
                        "timestamp": s["timestamp"], "duration": s["duration"], "text": s["text"], "link": s["link"]})
        buf = io.BytesIO(); buf.write(json.dumps(result, ensure_ascii=False, indent=2).encode()); buf.seek(0)
        return send_file(buf, mimetype="application/json", as_attachment=True, download_name=f"{filename}.json")
    else:
        buf = io.StringIO(); buf.write("\ufeff")
        writer = csv.writer(buf)
        if structure == "per_video":
            for v in videos:
                writer.writerow([f"=== {v['title']} | https://www.youtube.com/watch?v={v['video_id']} ==="])
                writer.writerow(["Timestamp", "Duration", "Transcript", "Video Link"])
                for s in v["segments"]:
                    writer.writerow([s["timestamp"], s["duration"], s["text"], s["link"]])
                writer.writerow([])
        else:
            writer.writerow(["Video Title", "Video ID", "Timestamp", "Duration", "Transcript", "Video Link"])
            for v in videos:
                for s in v["segments"]:
                    writer.writerow([v["title"], v["video_id"], s["timestamp"], s["duration"], s["text"], s["link"]])
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode("utf-8")), mimetype="text/csv", as_attachment=True, download_name=f"{filename}.csv")



# ═══════════════════════════════════════
#  翻译 API
# ═══════════════════════════════════════

@app.route("/api/translate", methods=["POST"])
def api_translate():
    """批量翻译文本"""
    data = request.get_json()
    texts = data.get("texts", [])
    target = data.get("target", "zh-CN")
    if not texts:
        return jsonify({"error": "没有待翻译文本"}), 400
    try:
        translator = GoogleTranslator(source="auto", target=target)
        results = []
        for text in texts:
            if not text or not text.strip():
                results.append("")
                continue
            try:
                # 限制单条长度避免超时
                chunk = text[:1500]
                results.append(translator.translate(chunk))
            except Exception:
                results.append(text)  # 翻译失败保留原文
        return jsonify({"translations": results, "target": target})
    except Exception as e:
        return jsonify({"error": f"翻译服务不可用: {str(e)}"}), 500


# ═══════════════════════════════════════
#  主页面
# ═══════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


# ── 共享：拉取单个视频元信息 ──
def _fetch_video_info(vid: str, user_key: str = None) -> dict | None:
    """拉取视频元信息，失败返回 None"""
    try:
        vdata = _youtube_get(f"videos?part=snippet,statistics,contentDetails&id={vid}",
                             key_override=user_key, retries=2)
    except Exception as e:
        logger.warning(f"video_info: failed for {vid}: {e}")
        return None
    items = vdata.get("items", [])
    if not items:
        return None
    info = items[0]
    snippet = info["snippet"]
    stats = info.get("statistics", {})
    cd = info.get("contentDetails", {})
    return {
        "video_id": vid,
        "title": snippet["title"],
        "channel": snippet["channelTitle"],
        "published_at": snippet["publishedAt"],
        "views": int(stats.get("viewCount", 0)),
        "url": f"https://www.youtube.com/watch?v={vid}",
        "duration": _format_duration(cd.get("duration", "")),
        "has_captions": cd.get("caption", "false") == "true",
        "comment_count": int(stats.get("commentCount", 0)),
    }

# ── 共享：拉取视频评论 ──
def _fetch_video_comments(vid: str, videos: list, min_likes: int = 0,
                          user_key: str = None, fetch_transcripts: bool = False) -> int:
    """拉取单个视频的评论+可选字幕，追加到 videos 列表。返回实际添加的视频数(0或1)"""
    info = _fetch_video_info(vid, user_key)
    if not info:
        return 0

    comments = _fetch_comments(vid, min_likes=min_likes, key_override=user_key)
    if not comments and not fetch_transcripts:
        return 0

    video = {**info, "comments": comments, "transcripts": []}

    if fetch_transcripts:
        captions = _extract_captions(vid)
        if captions:
            for seg in captions["segments"]:
                ss = int(seg["start"])
                mm, ss2 = divmod(ss, 60)
                video["transcripts"].append({
                    "timestamp": f"{mm:02d}:{ss2:02d}",
                    "duration": round(seg["duration"], 1),
                    "text": seg["text"],
                    "link": f"https://www.youtube.com/watch?v={vid}&t={ss}",
                })

    if video["comments"] or video["transcripts"]:
        videos.append(video)
        return 1
    return 0

@app.route("/api/search")
def api_search():
    keyword = request.args.get("q", "").strip()
    page_token = request.args.get("pageToken", "")
    user_key = request.args.get("key", "").strip()
    if not keyword:
        return jsonify({"error": "关键词不能为空"}), 400

    path = f"search?part=snippet&order=date&q={keyword}&type=video&maxResults=50"
    if page_token:
        path += f"&pageToken={page_token}"

    key_override = user_key or None
    try:
        data = _youtube_get(path, key_override=key_override)
    except RuntimeError as e:
        err_msg = str(e)
        is_quota = "配额" in err_msg or "quota" in err_msg.lower()
        return jsonify({
            "error": err_msg,
            "quota_exhausted": is_quota,
        }), (429 if is_quota else 500)
    except Exception as e:
        return jsonify({"error": f"搜索失败: {str(e)}"}), 500

    videos = []
    video_ids = []
    for item in data.get("items", []):
        if item.get("id", {}).get("kind") != "youtube#video":
            continue
        s = item["snippet"]
        vid = item["id"]["videoId"]
        video_ids.append(vid)
        videos.append({
            "video_id": vid,
            "title": s["title"],
            "channel": s["channelTitle"],
            "published_at": s["publishedAt"],
            "thumbnail": s["thumbnails"]["default"]["url"],
        })

    if not video_ids:
        return jsonify({"videos": []})

    try:
        stats_data = _youtube_get(
            f"videos?part=statistics,contentDetails&id={','.join(video_ids)}",
            key_override=user_key or None
        )
    except Exception as e:
        return jsonify({"error": f"统计失败: {str(e)}"}), 500

    stats_map = {}
    for item in stats_data.get("items", []):
        s = item.get("statistics", {})
        cd = item.get("contentDetails", {})
        stats_map[item["id"]] = {
            "view_count": int(s.get("viewCount", 0)),
            "comment_count": int(s.get("commentCount", 0)),
            "duration": _format_duration(cd.get("duration", "")),
            "has_captions": cd.get("caption", "false") == "true",
        }

    for v in videos:
        st = stats_map.get(v["video_id"], {})
        v["view_count"] = st.get("view_count", 0)
        v["comment_count"] = st.get("comment_count", 0)
        v["duration"] = st.get("duration", "")
        v["has_captions"] = st.get("has_captions", False)

    videos.sort(key=lambda v: v["view_count"], reverse=True)

    return jsonify({
        "videos": videos,
        "nextPageToken": data.get("nextPageToken", ""),
    })


def _fetch_comments(vid: str, min_likes: int = 0, max_pages: int = 5, key_override: str = None) -> list[dict]:
    """拉取评论，支持翻页（每页100条，最多5页=500条），含重试"""
    comments = []
    page_token = ""
    for page_num in range(max_pages):
        try:
            path = f"commentThreads?part=snippet&videoId={vid}&maxResults=100&order=time"
            if page_token:
                path += f"&pageToken={page_token}"
            cdata = _youtube_get(path, key_override=key_override, retries=2)
        except Exception:
            if page_num == 0:
                raise  # 第一页失败就抛出去，别静默吞掉
            break  # 后续页失败则停止翻页，保留已有数据
        for item in cdata.get("items", []):
            top = item["snippet"]["topLevelComment"]["snippet"]
            likes = top.get("likeCount", 0)
            if likes < min_likes:
                continue
            comments.append({
                "author": top["authorDisplayName"],
                "text": top["textDisplay"],
                "likes": likes,
                "published_at": top["publishedAt"],
            })
        page_token = cdata.get("nextPageToken", "")
        if not page_token:
            break
    return comments


@app.route("/api/preview", methods=["POST"])
def api_preview():
    data = request.get_json()
    video_ids = data.get("video_ids", [])
    if not video_ids:
        return jsonify({"error": "参数错误"}), 400
    comments = _fetch_comments(video_ids[0])
    return jsonify({"comments": comments, "total": len(comments)})


@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json()
    video_ids = data.get("video_ids", [])
    fmt = data.get("format", "csv")
    min_likes = int(data.get("min_likes", 0))
    structure = data.get("structure", "flat")
    user_key = data.get("api_key", "").strip() or None

    if not video_ids:
        return _err("请选择至少一个视频")

    # ── 按视频拉取 ──
    videos = []
    stats = {"total_comments": 0, "total_likes": 0, "video_count": 0}

    for vid in video_ids:
        try:
            vdata = _youtube_get(f"videos?part=snippet,statistics&id={vid}", key_override=user_key, retries=2)
        except Exception as e:
            print(f"[export] video info failed for {vid}: {e}")
            continue
        items = vdata.get("items", [])
        if not items:
            continue
        info = items[0]
        title = info["snippet"]["title"]
        channel = info["snippet"]["channelTitle"]
        pub = info["snippet"]["publishedAt"]
        views = info["statistics"].get("viewCount", 0)
        url = f"https://www.youtube.com/watch?v={vid}"

        comments = _fetch_comments(vid, min_likes=min_likes, key_override=user_key)
        if not comments:
            continue

        stats["video_count"] += 1
        video = {
            "video_id": vid, "title": title, "channel": channel,
            "published_at": pub, "views": views, "url": url,
            "comments": comments,
        }
        videos.append(video)
        for c in comments:
            stats["total_comments"] += 1
            stats["total_likes"] += c["likes"]

    if not videos:
        is_quota = _quota_exhausted
        return jsonify({
            "error": "所选视频暂无数据",
            "detail": "可能原因：API 配额用完、视频评论已关闭、或网络异常，请稍后重试。",
            "quota_exhausted": is_quota,
        }), (429 if is_quota else 404)

    date_str = datetime.now().strftime("%Y-%m-%d")
    safe_title = "".join(c if c.isalnum() or c in "-_" else "" for c in videos[0]["title"][:30]).strip()[:30]
    filename = f"{safe_title}-Comments-{date_str}" if safe_title else f"comments_{date_str}"

    if fmt == "xlsx":
        return _export_comments_xlsx(videos, filename, structure)
    elif fmt == "json":
        return _export_comments_json(videos, filename, structure)
    else:
        return _export_comments_csv(videos, filename, structure)


def _export_comments_csv(videos, filename, structure):
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = csv.writer(buf)

    if structure == "per_video":
        for v in videos:
            writer.writerow([f"=== {v['title']} | {v['url']} | {v['channel']} ==="])
            writer.writerow(["评论作者", "评论内容", "点赞数", "评论时间"])
            for c in v["comments"]:
                writer.writerow([c["author"], c["text"], c["likes"], c["published_at"]])
            writer.writerow([])
    else:
        writer.writerow(["评论作者", "评论内容", "点赞数", "评论时间",
                          "视频标题", "视频链接", "频道", "视频播放量", "视频发布时间"])
        for v in videos:
            for c in v["comments"]:
                writer.writerow([c["author"], c["text"], c["likes"], c["published_at"],
                                  v["title"], v["url"], v["channel"], v["views"], v["published_at"]])

    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8")), mimetype="text/csv",
                     as_attachment=True, download_name=f"{filename}.csv")


def _export_comments_xlsx(videos, filename, structure):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({"error": "Excel 导出需要 openpyxl 库"}), 500

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")

    if structure == "per_video":
        # 每个视频一个 Sheet
        first = True
        for v in videos:
            name = v["title"][:31].replace("/","-").replace("\\","-")
            if first:
                ws = wb.active; ws.title = name; first = False
            else:
                ws = wb.create_sheet(name)
            headers = ["评论作者", "评论内容", "点赞数", "评论时间"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = hf; cell.fill = hfill
            for r, c in enumerate(v["comments"], 2):
                for ci, val in enumerate([c["author"], c["text"], c["likes"], c["published_at"]], 1):
                    ws.cell(row=r, column=ci, value=val)
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 18
    else:
        # 统一汇总
        ws = wb.active; ws.title = "评论数据"
        headers = ["评论作者", "评论内容", "点赞数", "评论时间",
                   "视频标题", "视频链接", "频道", "视频播放量", "视频发布时间"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hf; cell.fill = hfill
        r = 2
        for v in videos:
            for c in v["comments"]:
                for ci, val in enumerate([c["author"], c["text"], c["likes"], c["published_at"],
                                           v["title"], v["url"], v["channel"], v["views"], v["published_at"]], 1):
                    ws.cell(row=r, column=ci, value=val)
                r += 1
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 18

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"{filename}.xlsx")


def _export_comments_json(videos, filename, structure):
    if structure == "per_video":
        result = {"exported_at": datetime.now().isoformat(), "videos": []}
        for v in videos:
            result["videos"].append({
                "title": v["title"], "url": v["url"], "channel": v["channel"],
                "views": v["views"], "published_at": v["published_at"],
                "comment_count": len(v["comments"]),
                "comments": [{"author": c["author"], "text": c["text"],
                               "likes": c["likes"], "published_at": c["published_at"]} for c in v["comments"]],
            })
    else:
        result = {"exported_at": datetime.now().isoformat(), "comments": []}
        for v in videos:
            for c in v["comments"]:
                result["comments"].append({
                    "author": c["author"], "text": c["text"], "likes": c["likes"],
                    "published_at": c["published_at"], "video_title": v["title"],
                    "video_url": v["url"], "channel": v["channel"],
                    "video_views": v["views"], "video_published": v["published_at"],
                })
    buf = io.BytesIO()
    buf.write(json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.seek(0)
    return send_file(buf, mimetype="application/json", as_attachment=True, download_name=f"{filename}.json")


@app.route("/api/export/stats", methods=["POST"])
def api_export_stats():
    data = request.get_json()
    video_ids = data.get("video_ids", [])
    min_likes = int(data.get("min_likes", 0))
    total_comments = 0
    total_likes = 0
    video_count = 0
    for vid in video_ids:
        comments = _fetch_comments(vid, min_likes=min_likes)
        if comments:
            video_count += 1
            total_comments += len(comments)
            total_likes += sum(c["likes"] for c in comments)
    return jsonify({
        "video_count": video_count,
        "total_comments": total_comments,
        "total_likes": total_likes,
    })

@app.route("/api/export/combined", methods=["POST"])
def api_export_combined():
    """合并导出：评论 + 字幕，按视频分组，可选翻译"""
    data = request.get_json()
    video_ids = data.get("video_ids", [])
    fmt = data.get("format", "csv")
    min_likes = int(data.get("min_likes", 0))
    translate_to = data.get("translate_to", "")
    user_key = data.get("api_key", "").strip() or None

    if not video_ids:
        return _err("请选择至少一个视频")

    # ── 按视频拉取评论 + 字幕 ──
    videos = []
    for vid in video_ids:
        try:
            vdata = _youtube_get(f"videos?part=snippet,statistics&id={vid}", key_override=user_key, retries=2)
        except Exception as e:
            print(f"[export] video info failed for {vid}: {e}")
            continue
        items = vdata.get("items", [])
        if not items:
            continue
        info = items[0]
        title = info["snippet"]["title"]
        channel = info["snippet"]["channelTitle"]
        pub = info["snippet"]["publishedAt"]
        views = info["statistics"].get("viewCount", 0)
        url = f"https://www.youtube.com/watch?v={vid}"

        video = {
            "video_id": vid,
            "title": title,
            "channel": channel,
            "published_at": pub,
            "views": views,
            "url": url,
            "comments": [],
            "transcripts": [],
        }

        # 评论
        for c in _fetch_comments(vid, min_likes=min_likes):
            video["comments"].append(c)

        # 字幕
        result = _extract_captions(vid)
        if result:
            for seg in result["segments"]:
                ss = int(seg["start"])
                mm = ss // 60
                ss2 = ss % 60
                ts = f"{mm:02d}:{ss2:02d}"
                link = f"https://www.youtube.com/watch?v={vid}&t={ss}"
                video["transcripts"].append({
                    "timestamp": ts,
                    "duration": round(seg["duration"], 1),
                    "text": seg["text"],
                    "link": link,
                })

        if video["comments"] or video["transcripts"]:
            videos.append(video)

    # ── 翻译 ──
    if translate_to and videos and _HAS_TRANSLATOR:
        try:
            translator = GoogleTranslator(source="auto", target=translate_to)
            for v in videos:
                for c in v.get("comments", []):
                    try:
                        c["text_translated"] = translator.translate(c["text"][:1500])
                    except Exception:
                        c["text_translated"] = c["text"]
                for t in v.get("transcripts", []):
                    try:
                        t["text_translated"] = translator.translate(t["text"][:1500])
                    except Exception:
                        t["text_translated"] = t["text"]
        except Exception:
            pass  # 翻译失败不影响导出

    if not videos:
        return jsonify({"error": "所选视频暂无数据"}), 404

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"youtube_data_{date_str}"

    structure = data.get("structure", "flat")
    if fmt == "json":
        return _export_combined_json(videos, filename)
    elif fmt == "xlsx":
        return _export_combined_xlsx(videos, filename, structure)
    else:
        return _export_combined_csv(videos, filename, structure)


def _export_combined_csv(videos, filename, structure="flat"):
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = csv.writer(buf)

    if structure == "per_video":
        for v in videos:
            views_str = str(v['views']); writer.writerow([f"=== {v['title']} | {v['url']} | {v['channel']} | {views_str} 次播放 ==="])
            if v["comments"]:
                writer.writerow(["--- 评论 ({0} 条) ---".format(len(v["comments"]))])
                writer.writerow(["评论作者", "评论内容", "点赞数", "评论时间"])
                for c in v["comments"]:
                    writer.writerow([c["author"], c["text"], c["likes"], c["published_at"]])
            if v["transcripts"]:
                writer.writerow(["--- 字幕 ({0} 段) ---".format(len(v["transcripts"]))])
                writer.writerow(["时间戳", "时长", "字幕内容"])
                for t in v["transcripts"]:
                    writer.writerow([t["timestamp"], t["duration"], t["text"]])
            writer.writerow([])
    else:
        has_trans = any(c.get("text_translated") for v in videos for c in v.get("comments",[])) or any(t.get("text_translated") for v in videos for t in v.get("transcripts",[]))
        headers = ["视频标题", "视频链接", "频道", "视频播放量", "视频发布时间",
                    "类型", "作者/时间戳", "内容", "点赞/时长", "发布时间"]
        if has_trans:
            headers.insert(8, "翻译")
        writer.writerow(headers)
        for v in videos:
            for c in v["comments"]:
                row = [v["title"], v["url"], v["channel"], v["views"], v["published_at"],
                        "评论", c["author"], c["text"], c["likes"], c["published_at"]]
                if has_trans:
                    row.insert(8, c.get("text_translated", ""))
                writer.writerow(row)
            for t in v["transcripts"]:
                row = [v["title"], v["url"], v["channel"], v["views"], v["published_at"],
                        "字幕", t["timestamp"], t["text"], t["duration"], ""]
                if has_trans:
                    row.insert(8, t.get("text_translated", ""))
                writer.writerow(row)

    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8")), mimetype="text/csv",
                     as_attachment=True, download_name=f"{filename}.csv")


def _export_combined_xlsx(videos, filename, structure="flat"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({"error": "Excel 导出需要 openpyxl 库"}), 500

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")

    if structure == "per_video":
        # 每个视频一个 Sheet，内含评论+字幕两个区域
        first = True
        for vi, v in enumerate(videos):
            name = v["title"][:28].replace("/","-").replace("\\","-")
            ws = wb.active if first else wb.create_sheet(name)
            if first: ws.title = name; first = False
            # 评论区域
            r = 1
            if v["comments"]:
                for col, h in enumerate(["评论作者", "评论内容", "点赞数", "评论时间"], 1):
                    cell = ws.cell(row=r, column=col, value=h); cell.font = hf; cell.fill = hfill
                r += 1
                for c in v["comments"]:
                    for ci, val in enumerate([c["author"], c["text"], c["likes"], c["published_at"]], 1):
                        ws.cell(row=r, column=ci, value=val)
                    r += 1
                r += 1  # spacer
            # 字幕区域
            if v["transcripts"]:
                for col, h in enumerate(["时间戳", "时长", "字幕内容", "链接"], 1):
                    cell = ws.cell(row=r, column=col, value=h); cell.font = hf; cell.fill = hfill
                r += 1
                for t in v["transcripts"]:
                    for ci, val in enumerate([t["timestamp"], t["duration"], t["text"], t["link"]], 1):
                        ws.cell(row=r, column=ci, value=val)
                    r += 1
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 18
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"{filename}.xlsx")

    # ── 汇总 Sheet（统一扁平）──
    ws0 = wb.active
    ws0.title = "汇总"
    u_headers = ["视频标题", "视频链接", "频道", "视频播放量", "视频发布时间",
                  "类型", "作者/时间戳", "内容", "点赞/时长", "发布时间"]
    for col, h in enumerate(u_headers, 1):
        cell = ws0.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
    row_idx = 2
    for v in videos:
        for c in v["comments"]:
            for ci, val in enumerate([v["title"], v["url"], v["channel"], v["views"], v["published_at"],
                                       "评论", c["author"], c["text"], c["likes"], c["published_at"]], 1):
                ws0.cell(row=row_idx, column=ci, value=val)
            row_idx += 1
        for t in v["transcripts"]:
            for ci, val in enumerate([v["title"], v["url"], v["channel"], v["views"], v["published_at"],
                                       "字幕", t["timestamp"], t["text"], t["duration"], ""], 1):
                ws0.cell(row=row_idx, column=ci, value=val)
            row_idx += 1
    for col, w in enumerate([40, 40, 20, 14, 18, 8, 16, 60, 10, 18], 1):
        ws0.column_dimensions[chr(64+col) if col <= 26 else "A"].width = w

    # ── 评论 Sheet ──
    ws1 = wb.create_sheet("评论")
    c_headers = ["评论作者", "评论内容", "点赞数", "评论时间",
                  "视频标题", "视频链接", "频道", "视频播放量", "视频发布时间"]
    for col, h in enumerate(c_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
    r = 2
    for v in videos:
        for c in v["comments"]:
            for ci, val in enumerate([c["author"], c["text"], c["likes"], c["published_at"],
                                       v["title"], v["url"], v["channel"], v["views"], v["published_at"]], 1):
                ws1.cell(row=r, column=ci, value=val)
            r += 1
    for col, w in enumerate([18, 60, 10, 18, 40, 40, 20, 14, 18], 1):
        ws1.column_dimensions[chr(64+col)].width = w

    # ── 字幕 Sheet ──
    ws2 = wb.create_sheet("字幕")
    t_headers = ["视频标题", "视频ID", "时间戳", "时长(秒)", "字幕内容", "视频链接"]
    for col, h in enumerate(t_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
    r = 2
    for v in videos:
        for t in v["transcripts"]:
            for ci, val in enumerate([v["title"], v["video_id"], t["timestamp"],
                                       t["duration"], t["text"], t["link"]], 1):
                ws2.cell(row=r, column=ci, value=val)
            r += 1
    for col, w in enumerate([40, 14, 10, 8, 60, 40], 1):
        ws2.column_dimensions[chr(64+col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"{filename}.xlsx")


def _export_combined_json(videos, filename):
    """按视频分组的结构化 JSON"""
    result = {
        "exported_at": datetime.now().isoformat(),
        "total_videos": len(videos),
        "videos": [],
    }
    for v in videos:
        result["videos"].append({
            "title": v["title"],
            "url": v["url"],
            "channel": v["channel"],
            "views": v["views"],
            "published_at": v["published_at"],
            "comment_count": len(v["comments"]),
            "transcript_segments": len(v["transcripts"]),
            "comments": [{
                "author": c["author"],
                "text": c["text"],
                "likes": c["likes"],
                "published_at": c["published_at"],
            } for c in v["comments"]],
            "transcripts": [{
                "timestamp": t["timestamp"],
                "duration": t["duration"],
                "text": t["text"],
                "link": t["link"],
            } for t in v["transcripts"]],
        })
    buf = io.BytesIO()
    buf.write(json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.seek(0)
    return send_file(buf, mimetype="application/json", as_attachment=True,
                     download_name=f"{filename}.json")


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
